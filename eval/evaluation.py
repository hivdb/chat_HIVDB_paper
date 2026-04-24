#!/usr/bin/env python3
"""Evaluate model outputs against human answers with lenient matching."""

from __future__ import annotations

import argparse
import logging
import sys
from pathlib import Path
from typing import List, Tuple, Iterable

import pandas as pd

ROOT = Path(__file__).resolve().parent
if str(ROOT.parent) not in sys.path:
    sys.path.append(str(ROOT.parent))

from eval import config  # type: ignore
from eval import statistics as stat_utils  # type: ignore
from eval.plots import generate_figures, plot_metric_by_qid  # type: ignore
from eval.scoring import build_detail_rows, ensure_norm, evaluate_group, evaluate_model, load_dataset  # type: ignore
# Statistical helpers
from scipy.stats import fisher_exact
import numpy as np

from eval.normalize import human_answer_counts, slugify
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

FAMILY_COMPARISONS = {
    "GPT-4o": {
        "base": "GPT-4o base",
        "targets": ["GPT-4o FT", "GPT-4o FT+QSP", "GPT-4o QSP"],
    },
    "Llama3.1-70B": {
        "base": "Llama3.1-70B base",
        "targets": ["Llama3.1-70B FT", "Llama3.1-70B FT+QSP", "Llama3.1-70B QSP"],
    },
    "Llama3.1-8B": {
        "base": "Llama3.1-8B base",
        "targets": ["Llama3.1-8B FT", "Llama3.1-8B FT+QSP", "Llama3.1-8B QSP"],
    },
}
CONFIDENCE_LEVEL = 0.95
BOOTSTRAP_ITERATIONS = 5000
BOOTSTRAP_SEED = 42


def build_qid_metrics(
    scenario_df: pd.DataFrame,
    scenario: dict,
    norm_lookup: dict[str, str],
) -> List[dict]:
    rows: List[dict] = []
    if scenario_df.empty:
        return rows
    allow_partial = scenario.get("allow_partial_list", False)
    ref_col = config.REF_COL
    ref_norm = norm_lookup.get(ref_col)
    if not ref_norm:
        return rows
    for qid, qid_df in scenario_df.groupby("QID"):
        q_type = qid_df.get("Type", pd.Series([""])).iloc[0]
        question = qid_df.get("Question", pd.Series([""])).iloc[0]
        for model in scenario["models"]:
            pred_norm = norm_lookup.get(model)
            if not pred_norm or model not in qid_df.columns:
                continue
            metrics = evaluate_model(
                qid_df,
                model,
                ref_col,
                pred_norm,
                ref_norm,
                allow_partial_list=allow_partial,
            )
            metrics.update(
                {
                    "model": model,
                    "QID": qid,
                    "Type": q_type,
                    "Question": question,
                }
            )
            rows.append(metrics)
    return rows


def run(limit: int | None) -> Tuple[
    pd.DataFrame,
    List[dict],
    List[Tuple[str, str, pd.DataFrame]],
    List[dict],
    dict[str, pd.DataFrame],
]:
    df = load_dataset()
    if limit:
        df = df.head(limit)
    config.OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    scenario_frames: List[pd.DataFrame] = []
    detail_rows: List[dict] = []
    figure_specs: List[Tuple[str, str, pd.DataFrame]] = []
    qid_metrics_rows: List[dict] = []

    cache: dict = {}
    scenario_results: dict[str, pd.DataFrame] = {}
    scenario_qid_frames: dict[str, pd.DataFrame] = {}
    for scenario in config.SCENARIOS:
        scenario_df = df
        if filter_type := scenario.get("filter_type"):
            scenario_df = df[df["Type"] == filter_type].copy()
        if scenario_df.empty:
            continue

        missing_models = [model for model in scenario["models"] if model not in df.columns]
        if missing_models:
            logging.warning("Scenario '%s' missing models: %s", scenario["title"], ", ".join(missing_models))

        allow_partial = scenario.get("allow_partial_list", False)
        detail_type_filter = scenario.get("detail_types")
        detail_types = set(detail_type_filter) if detail_type_filter else None
        norm_lookup = {col: ensure_norm(df, col, cache) for col in [config.REF_COL, *scenario["models"]] if col in df.columns}
        subset = evaluate_group(
            scenario_df,
            scenario["models"],
            config.REF_COL,
            norm_lookup,
            allow_partial_list=allow_partial,
        )
        if subset.empty:
            continue
        scenario_id = scenario.get("scenario_id")
        if scenario.get("compare_to"):
            base_id = scenario.get("compare_to")
            base_subset = scenario_results.get(base_id)
            if base_subset is None:
                logging.warning("Scenario '%s' requires compare_to '%s', which is missing.", scenario["title"], base_id)
                continue
            metrics_cols = ["accuracy", "precision", "recall", "f1"]
            merged = subset.merge(
                base_subset[["model", *metrics_cols]],
                on="model",
                suffixes=("", "_base"),
            )
            for col in metrics_cols:
                merged[col] = merged[col] - merged[f"{col}_base"]
                merged.drop(columns=f"{col}_base", inplace=True)
            subset = merged
        if scenario_id:
            scenario_results[scenario_id] = subset.copy()
        if scenario.get("include_details", True):
            detail_rows.extend(
                build_detail_rows(
                    scenario_df,
                    scenario,
                    norm_lookup,
                    detail_types,
                )
            )
        scenario_qid = build_qid_metrics(
            scenario_df,
            scenario,
            norm_lookup,
        )
        qid_metrics_rows.extend(scenario_qid)
        scenario_qid_df = pd.DataFrame(scenario_qid)
        scenario_qid_frames[scenario["title"]] = scenario_qid_df
        scenario_frames.append(subset)
        def _format_title(title: str) -> str:
            if " - " in title:
                head, tail = title.split(" - ", 1)
                return f"{head} ({tail})"
            return title

        figure_specs.append((_format_title(scenario["title"]), scenario["title"], subset, scenario_qid_df))
    combined = pd.concat(scenario_frames, ignore_index=True) if scenario_frames else pd.DataFrame()
    return combined, detail_rows, figure_specs, qid_metrics_rows, scenario_qid_frames


def _write_excel(path: Path, sheets: dict[str, pd.DataFrame]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for name, df in sheets.items():
            df.to_excel(writer, sheet_name=name, index=False)


def _dataset_label_from_suffix(suffix: str) -> str:
    normalized = suffix.strip("_").lower()
    if not normalized:
        return "Full 150"
    if "full150" in normalized:
        return "Full 150"
    if "new30" in normalized:
        return "New 30"
    if "original120" in normalized or "orig120" in normalized:
        return "Original 120"
    return normalized.replace("_", " ").title()


def _aggregate_fisher_summary(
    qid_df: pd.DataFrame,
    comparisons: dict,
    fisher_metrics: list[str],
) -> pd.DataFrame:
    """Compute aggregated Fisher exact tests per family/target across all QIDs."""
    if qid_df is None or qid_df.empty:
        return pd.DataFrame()

    def _sum_counts(model: str) -> tuple[int, int, int, int] | None:
        subset = qid_df[qid_df["model"] == model]
        if subset.empty or not {"tp", "fp", "tn", "fn"}.issubset(subset.columns):
            return None
        totals = subset[["tp", "fp", "tn", "fn"]].sum()
        return int(totals["tp"]), int(totals["fp"]), int(totals["tn"]), int(totals["fn"])

    rows: list[dict] = []
    for family, mapping in comparisons.items():
        base_model = mapping.get("base")
        if not base_model:
            continue
        base_counts = _sum_counts(base_model)
        if base_counts is None:
            continue
        for target_model in mapping.get("targets", []):
            target_counts = _sum_counts(target_model)
            if target_counts is None:
                continue
            base_tp, base_fp, base_tn, base_fn = base_counts
            target_tp, target_fp, target_tn, target_fn = target_counts
            for metric in fisher_metrics:
                if metric == "accuracy":
                    base_pos, base_neg = base_tp + base_tn, base_fp + base_fn
                    target_pos, target_neg = target_tp + target_tn, target_fp + target_fn
                elif metric == "precision":
                    base_pos, base_neg = base_tp, base_fp
                    target_pos, target_neg = target_tp, target_fp
                elif metric == "recall":
                    base_pos, base_neg = base_tp, base_fn
                    target_pos, target_neg = target_tp, target_fn
                elif metric == "f1":
                    base_pos, base_neg = 2 * base_tp, base_fp + base_fn
                    target_pos, target_neg = 2 * target_tp, target_fp + target_fn
                else:
                    continue
                table = np.array([[base_pos, base_neg], [target_pos, target_neg]])
                try:
                    _, p_value = fisher_exact(table)
                except ValueError:
                    p_value = np.nan
                rows.append(
                    {
                        "family": family,
                        "comparison": target_model.replace(f"{family} ", ""),
                        "base_model": base_model,
                        "target_model": target_model,
                        "metric": metric,
                        "p_value": float(p_value),
                    }
                )
    summary = pd.DataFrame(rows)
    if summary.empty:
        return summary
    # Adjust p-values within each metric
    summary["adj_p_value"] = np.nan
    for metric in fisher_metrics:
        mask = summary["metric"] == metric
        metric_rows = summary.loc[mask]
        adjusted = stat_utils.benjamini_hochberg(metric_rows["p_value"].tolist())
        summary.loc[mask, "adj_p_value"] = adjusted
    return summary


def _metrics_from_count_arrays(
    tp: np.ndarray,
    tn: np.ndarray,
    fp: np.ndarray,
    fn: np.ndarray,
) -> dict[str, np.ndarray]:
    total = tp + tn + fp + fn
    accuracy = np.divide(tp + tn, total, out=np.zeros_like(total, dtype=float), where=total != 0)
    precision_den = tp + fp
    precision = np.divide(tp, precision_den, out=np.zeros_like(tp, dtype=float), where=precision_den != 0)
    recall_den = tp + fn
    recall = np.divide(tp, recall_den, out=np.zeros_like(tp, dtype=float), where=recall_den != 0)
    f1_den = precision + recall
    f1 = np.divide(2 * precision * recall, f1_den, out=np.zeros_like(precision, dtype=float), where=f1_den != 0)
    return {
        "accuracy": accuracy,
        "precision": precision,
        "recall": recall,
        "f1": f1,
    }


def build_bar_chart_confidence_intervals(details_path: Path, subset: pd.DataFrame) -> pd.DataFrame:
    if subset.empty:
        return pd.DataFrame(columns=["model", "display_label", "metric", "value", "ci_low", "ci_high", "confidence_level"])
    details_df = pd.read_excel(details_path, sheet_name="All", keep_default_na=False, na_filter=False)
    cache: dict[str, str] = {}
    ref_norm = ensure_norm(details_df, config.REF_COL, cache)
    alpha = 1.0 - CONFIDENCE_LEVEL
    lower_pct = 100 * (alpha / 2)
    upper_pct = 100 * (1 - alpha / 2)
    rng = np.random.default_rng(BOOTSTRAP_SEED)
    rows: list[dict[str, float | str]] = []

    ordered_subset = subset.copy()
    if "display_order" in ordered_subset.columns:
        ordered_subset = ordered_subset.sort_values("display_order")

    for item in ordered_subset.itertuples(index=False):
        model = str(getattr(item, "model"))
        answer_col = f"{model} Answer"
        if answer_col not in details_df.columns:
            raise ValueError(f"Answer column missing from detailed evaluation workbook: {answer_col}")
        pred_norm = ensure_norm(details_df, answer_col, cache)

        row_counts: list[tuple[int, int, int, int]] = []
        for _, detail_row in details_df.iterrows():
            allow_partial = str(detail_row.get("Type", "")).strip().lower() == "list"
            counts, _ = human_answer_counts(
                str(detail_row.get("Type", "")),
                str(detail_row.get(pred_norm, "")),
                str(detail_row.get(ref_norm, "")),
                question_text=str(detail_row.get("Question", "")),
                ref_raw=str(detail_row.get(config.REF_COL, "")),
                pred_raw=str(detail_row.get(answer_col, "")),
                allow_partial_list=allow_partial,
            )
            row_counts.append((counts["tp"], counts["tn"], counts["fp"], counts["fn"]))

        count_array = np.asarray(row_counts, dtype=int)
        tp = count_array[:, 0]
        tn = count_array[:, 1]
        fp = count_array[:, 2]
        fn = count_array[:, 3]
        point_metrics = _metrics_from_count_arrays(
            np.array([tp.sum()]),
            np.array([tn.sum()]),
            np.array([fp.sum()]),
            np.array([fn.sum()]),
        )

        sample_size = len(count_array)
        bootstrap_idx = rng.integers(0, sample_size, size=(BOOTSTRAP_ITERATIONS, sample_size))
        boot_tp = tp[bootstrap_idx].sum(axis=1)
        boot_tn = tn[bootstrap_idx].sum(axis=1)
        boot_fp = fp[bootstrap_idx].sum(axis=1)
        boot_fn = fn[bootstrap_idx].sum(axis=1)
        boot_metrics = _metrics_from_count_arrays(boot_tp, boot_tn, boot_fp, boot_fn)

        display_label = str(getattr(item, "display_label", model))
        for metric in ["accuracy", "precision", "recall", "f1"]:
            rows.append(
                {
                    "model": model,
                    "display_label": display_label,
                    "metric": metric,
                    "value": float(point_metrics[metric][0]) * 100.0,
                    "ci_low": float(np.percentile(boot_metrics[metric], lower_pct)) * 100.0,
                    "ci_high": float(np.percentile(boot_metrics[metric], upper_pct)) * 100.0,
                    "confidence_level": CONFIDENCE_LEVEL,
                }
            )

    columns = ["model", "display_label", "metric", "value", "ci_low", "ci_high", "confidence_level"]
    return pd.DataFrame(rows)[columns]


def _build_fisher_qid_sheet(fisher: pd.DataFrame) -> pd.DataFrame:
    """Expand fisher results into a long per-QID dataframe."""
    if fisher is None or fisher.empty:
        return pd.DataFrame()
    records: list[dict] = []
    for _, row in fisher.iterrows():
        family = row.get("family")
        comparison = row.get("comparison")
        metric = row.get("metric")
        test_name = row.get("test")
        for col in fisher.columns:
            if not col.startswith("p_value_qid_"):
                continue
            qid = int(col.split("_")[-1])
            adj_col = f"adj_p_qid_{qid}"
            base_col = f"base_qid_{qid}"
            target_col = f"target_qid_{qid}"
            records.append(
                {
                    "family": family,
                    "comparison": comparison,
                    "metric": metric,
                    "test": test_name,
                    "QID": qid,
                    "base": row.get(base_col),
                    "target": row.get(target_col),
                    "p_value": row.get(col),
                    "adj_p": row.get(adj_col),
                }
            )
    return pd.DataFrame(records)


def _drop_qid_cols(df: pd.DataFrame) -> pd.DataFrame:
    """Remove per-QID columns for compact summary output."""
    return df[
        [
            c
            for c in df.columns
            if not (
                c.startswith("base_qid_")
                or c.startswith("target_qid_")
                or c.startswith("p_value_qid_")
                or c.startswith("adj_p_qid_")
            )
        ]
    ].copy()


def _build_table3(
    fisher_qid_df: pd.DataFrame,
    qid_metrics_df: pd.DataFrame | None,
) -> pd.DataFrame:
    """Construct Table3: precision/recall with significance markers for select QIDs."""
    if fisher_qid_df is None or fisher_qid_df.empty or qid_metrics_df is None or qid_metrics_df.empty:
        return pd.DataFrame()

    qid_question = qid_metrics_df.groupby("QID")["Question"].first().to_dict()

    # Identify where any metric is significant (p<0.05 and target > base) per family/target/QID
    sig_map: dict[tuple[str, str, int], set[str]] = {}
    sig_any: set[tuple[str, int]] = set()
    sig_info: dict[tuple[str, str, str, int], tuple[float, float]] = {}
    for _, row in fisher_qid_df.iterrows():
        family = row.get("family")
        comparison = row.get("comparison")
        metric = row.get("metric")
        # Only consider the metrics that are displayed (precision/recall)
        if metric not in {"precision", "recall"}:
            continue
        qid = row.get("QID")
        try:
            p_val = float(row.get("p_value"))
        except (TypeError, ValueError):
            continue
        try:
            base_val = float(row.get("base"))
            target_val = float(row.get("target"))
        except (TypeError, ValueError):
            continue
        try:
            adj_val = float(row.get("adj_p"))
        except (TypeError, ValueError):
            continue
        if comparison == "FT+QSP":
            continue
        if adj_val < 0.05 and target_val > base_val:
            sig_map.setdefault((str(family), str(comparison), int(qid)), set()).add(str(metric))
            sig_any.add((str(family), int(qid)))
            sig_info[(str(family), str(comparison), str(metric), int(qid))] = (
                p_val,
                adj_val,
            )

    def _metric_value(qid: int, model: str, metric: str) -> float | None:
        subset = qid_metrics_df[(qid_metrics_df["QID"] == qid) & (qid_metrics_df["model"] == model)]
        if subset.empty or metric not in subset:
            return None
        try:
            return float(subset[metric].iloc[0])
        except Exception:
            return None

    # Ordering of QIDs to keep the layout stable
    qid_order: dict[str, list[int]] = {}

    def _target_model(family: str, label: str) -> str | None:
        mapping = FAMILY_COMPARISONS.get(family, {})
        targets = mapping.get("targets", [])
        if label == "QSP":
            # Only use the pure QSP variant; do not fall back to FT+QSP.
            for target in targets:
                if target.endswith(" QSP") and "FT+QSP" not in target:
                    return target
            return None
        if label == "FT":
            # Only use FT (exclude FT+QSP combined variants).
            for target in targets:
                if "FT" in target and "FT+QSP" not in target and "QSP" not in target.split():
                    return target
            return None
        return None

    records: list[dict] = []
    families = sorted({fam for fam, _ in sig_any}, key=lambda f: {"GPT-4o": 0, "Llama3.1-70B": 1, "Llama3.1-8B": 2}.get(f, 99))
    for family in families:
        fam_sig = [qid for fam, qid in sig_any if fam == family]
        if not fam_sig:
            continue
        ordered_qids = sorted(fam_sig)
        base_model = FAMILY_COMPARISONS.get(family, {}).get("base")
        ft_model = _target_model(family, "FT")
        qsp_model = _target_model(family, "QSP")
        if not base_model:
            continue
        for qid in ordered_qids:
            if (family, qid) not in sig_any:
                continue
            # Skip rows with no significant (adj_p<0.05) entries
            if (family, qid) not in sig_any:
                continue
            row = {
                "Model": family,
                "QID": qid,
                "Question": qid_question.get(qid, ""),
                "base_prec": _metric_value(qid, base_model, "precision"),
                "FT_prec": "",
                "QSP_prec": "",
                "base_rec": _metric_value(qid, base_model, "recall"),
                "FT_rec": "",
                "QSP_rec": "",
            }
            # Evaluate FT/QSP with significance thresholds
            for target_label, model_label in [("FT", ft_model), ("QSP", qsp_model)]:
                if not model_label:
                    continue
                tgt_prec = _metric_value(qid, model_label, "precision")
                tgt_rec = _metric_value(qid, model_label, "recall")
                sig_metrics = sig_map.get((family, target_label, qid), set())
                for metric, tgt_val, field in [
                    ("precision", tgt_prec, f"{target_label}_prec"),
                    ("recall", tgt_rec, f"{target_label}_rec"),
                ]:
                    if tgt_val is None:
                        continue
                    suffix = ""
                    if metric in sig_metrics:
                        _, adj_val = sig_info.get((family, target_label, metric, qid), (None, None))
                        if adj_val is not None and not np.isnan(adj_val):
                            if adj_val < 0.01:
                                suffix = "**"
                            elif adj_val < 0.05:
                                suffix = "*"
                    row[field] = f"{tgt_val * 100:.1f}{suffix}"
            # Format base values as percentages
            if row["base_prec"] is not None:
                row["base_prec"] = f"{row['base_prec'] * 100:.1f}"
            else:
                row["base_prec"] = ""
            if row["base_rec"] is not None:
                row["base_rec"] = f"{row['base_rec'] * 100:.1f}"
            else:
                row["base_rec"] = ""
            records.append(row)
    df = pd.DataFrame(records, columns=["Model", "QID", "Question", "base_prec", "FT_prec", "QSP_prec", "base_rec", "FT_rec", "QSP_rec"])
    if df.empty:
        return df
    model_order = {"GPT-4o": 0, "Llama3.1-70B": 1, "Llama3.1-8B": 2}
    df["__order"] = df["Model"].map(model_order).fillna(99)
    df.sort_values(["__order", "QID"], inplace=True)
    df.drop(columns="__order", inplace=True)
    return df


def _model_for_comparison(family: str, comparison: str) -> str | None:
    mapping = FAMILY_COMPARISONS.get(family, {})
    targets = mapping.get("targets", [])
    if comparison == "FT":
        for target in targets:
            if "FT+QSP" in target or target.endswith(" QSP"):
                continue
            if target.endswith(" FT"):
                return target
    if comparison == "QSP":
        for target in targets:
            if target.endswith(" QSP") and "FT+QSP" not in target:
                return target
    if comparison == "FT+QSP":
        for target in targets:
            if "FT+QSP" in target:
                return target
    return None


def _build_suppfile2_individquest(
    qid_df: pd.DataFrame,
    fisher_qid_df: pd.DataFrame,
    comparisons: Iterable[str],
) -> pd.DataFrame:
    header = [
        "Question",
        "PcntTrue",
        "Model",
        "State",
        "TP",
        "FN",
        "FP",
        "TN",
        "Accuracy",
        "Precision",
        "Recall",
        "F1",
        "",
        "State",
        "TP",
        "FN",
        "FP",
        "TN",
        "Accuracy",
        "Precision",
        "Recall",
        "F1",
        "",
        "Accuracy P",
        "Rank",
        "Adjusted P",
        "BH sig",
        "",
        "Precision P",
        "Rank",
        "Adjusted P",
        "BH sig",
        "",
        "Recall P",
        "Rank",
        "Adjusted P",
        "BH sig",
        "",
        "Base (base model)",
    ]
    if qid_df is None or qid_df.empty:
        return pd.DataFrame(columns=header)

    qid_counts = qid_df.groupby("QID").first()
    pcnt_true = (qid_counts["tp"] + qid_counts["fn"]) / qid_counts["samples"]

    def _metrics(qid: int, model: str) -> dict[str, float]:
        row = qid_df[(qid_df["QID"] == qid) & (qid_df["model"] == model)]
        if row.empty:
            return {k: float("nan") for k in ["tp", "fn", "fp", "tn", "accuracy", "precision", "recall", "f1"]}
        row = row.iloc[0]
        return {
            "tp": float(row.get("tp", 0)),
            "fn": float(row.get("fn", 0)),
            "fp": float(row.get("fp", 0)),
            "tn": float(row.get("tn", 0)),
            "accuracy": float(row.get("accuracy", 0)),
            "precision": float(row.get("precision", 0)),
            "recall": float(row.get("recall", 0)),
            "f1": float(row.get("f1", 0)),
        }

    rows: list[list] = []
    qids = sorted(qid_df["QID"].unique())
    for family in ["GPT-4o", "Llama3.1-70B", "Llama3.1-8B"]:
        base_model = FAMILY_COMPARISONS.get(family, {}).get("base")
        if not base_model:
            continue
        for comparison in comparisons:
            target_model = _model_for_comparison(family, comparison)
            if not target_model:
                continue
            comp_df = fisher_qid_df[
                (fisher_qid_df["family"] == family)
                & (fisher_qid_df["comparison"] == comparison)
                & (fisher_qid_df["metric"].isin(["accuracy", "precision", "recall"]))
            ].copy()
            if comp_df.empty:
                continue
            comp_df["rank"] = comp_df.groupby("metric")["p_value"].rank(method="min")
            comp_df["bh_sig"] = comp_df["adj_p"].apply(lambda v: "yes" if pd.notna(v) and float(v) < 0.05 else "no")
            for qid in qids:
                base = _metrics(qid, base_model)
                target = _metrics(qid, target_model)
                row_vals = [None] * len(header)
                row_vals[0] = qid
                row_vals[1] = round(float(pcnt_true.get(qid, float("nan"))), 4)
                row_vals[2] = family
                row_vals[3] = "Base"
                row_vals[4:12] = [
                    base["tp"],
                    base["fn"],
                    base["fp"],
                    base["tn"],
                    round(base["accuracy"], 4),
                    round(base["precision"], 4),
                    round(base["recall"], 4),
                    round(base["f1"], 4),
                ]
                row_vals[13] = comparison
                row_vals[14:22] = [
                    target["tp"],
                    target["fn"],
                    target["fp"],
                    target["tn"],
                    round(target["accuracy"], 4),
                    round(target["precision"], 4),
                    round(target["recall"], 4),
                    round(target["f1"], 4),
                ]
                for metric, offset in [("accuracy", 23), ("precision", 28), ("recall", 33)]:
                    metric_row = comp_df[(comp_df["QID"] == qid) & (comp_df["metric"] == metric)]
                    if metric_row.empty:
                        continue
                    metric_row = metric_row.iloc[0]
                    row_vals[offset] = metric_row.get("p_value")
                    row_vals[offset + 1] = metric_row.get("rank")
                    row_vals[offset + 2] = metric_row.get("adj_p")
                    row_vals[offset + 3] = metric_row.get("bh_sig")
                rows.append(row_vals)
            rows.append([None] * len(header))
        rows.append([None] * len(header))

    legend = {
        0: "Base (base model)",
        1: "FT (fine-tuned model)",
        3: "P (p value)",
        4: "BH sig (Benjamini Hochberg): Is adjusted p value <0.05?",
        5: "Cyan: unadjusted p value <0.01; Light blue: unadjusted p value <0.05",
        8: "TP (true positive)",
        9: "TN (true negative)",
        10: "FP (false positive)",
        11: "FN (false negative)",
    }
    for idx, text in legend.items():
        if idx < len(rows):
            rows[idx][-1] = text
    return pd.DataFrame(rows, columns=header)


def _build_suppfile2_signedrank(qid_df: pd.DataFrame, paired_df: pd.DataFrame) -> pd.DataFrame:
    if qid_df is None or qid_df.empty:
        return pd.DataFrame()
    comparisons = ["Baseline", "FT", "QSP", "FT+QSP"]
    families = ["GPT-4o", "Llama3.1-70B", "Llama3.1-8B"]
    metrics = [("accuracy", "Accuracy"), ("precision", "Precision"), ("recall", "Recall"), ("f1", "F1")]
    block_starts = [1, 6, 11, 16]
    rows: list[list] = []

    def _row() -> list:
        return [None] * 20

    def _set_block(row, block_idx, label, values):
        start = block_starts[block_idx]
        row[start] = label
        for i, val in enumerate(values):
            row[start + 1 + i] = val

    def _model_value(qid, model, metric):
        row = qid_df[(qid_df["QID"] == qid) & (qid_df["model"] == model)]
        if row.empty:
            return float("nan")
        return float(row.iloc[0].get(metric, float("nan")))

    # Baseline header
    header = _row()
    header[0] = "Baseline"
    for idx, (_, metric_title) in enumerate(metrics):
        header[block_starts[idx]] = metric_title
    rows.append(header)
    header2 = _row()
    for idx, _ in enumerate(metrics):
        _set_block(header2, idx, "Question", families)
    rows.append(header2)

    qids = sorted(qid_df["QID"].unique())
    for qid in qids:
        row = _row()
        for idx, (metric, _) in enumerate(metrics):
            values = []
            for family in families:
                base_model = FAMILY_COMPARISONS.get(family, {}).get("base")
                val = _model_value(qid, base_model, metric)
                values.append(round(val, 4))
            _set_block(row, idx, qid, values)
        rows.append(row)
    rows.extend([_row(), _row()])

    for comparison in comparisons[1:]:
        header = _row()
        header[0] = comparison
        rows.append(header)
        header2 = _row()
        for idx, _ in enumerate(metrics):
            _set_block(header2, idx, "Question", families)
        rows.append(header2)
        for qid in qids:
            row = _row()
            for idx, (metric, _) in enumerate(metrics):
                values = []
                for family in families:
                    model = _model_for_comparison(family, comparison)
                    val = _model_value(qid, model, metric)
                    values.append(round(val, 4))
                _set_block(row, idx, qid, values)
            rows.append(row)
        for test_name, label in [("t-test", "P (T-Test)"), ("wilcoxon", "P (SignedRank)")]:
            row = _row()
            for idx, (metric, _) in enumerate(metrics):
                values = []
                for family in families:
                    match = paired_df[
                        (paired_df["family"] == family)
                        & (paired_df["comparison"] == comparison)
                        & (paired_df["metric"] == metric)
                        & (paired_df["test"] == test_name)
                    ]
                    values.append(float(match.iloc[0]["p_value"]) if not match.empty else float("nan"))
                _set_block(row, idx, label, values)
            rows.append(row)
        rows.extend([_row(), _row()])

    return pd.DataFrame(rows)


def _build_suppfile2_modelcomp_bh(paired_df: pd.DataFrame) -> pd.DataFrame:
    columns = [
        "Model 1",
        "Model 2",
        "Original Order",
        "Statistical test",
        "FDR rate",
        "",
        "Accuracy",
        "Rank",
        "adjust P",
        "BH sig",
        "",
        "Precision",
        "Rank",
        "adjust P",
        "BH sig",
        "",
        "Recall",
        "Rank",
        "adjust P",
        "BH sig",
        "",
        "F1",
        "Rank",
        "adjust P",
        "BH sig",
    ]
    if paired_df is None or paired_df.empty:
        return pd.DataFrame([columns])

    comparisons = ["FT", "QSP", "FT+QSP"]
    metrics = ["accuracy", "precision", "recall", "f1"]
    filtered = paired_df[paired_df["comparison"].isin(comparisons)].copy()
    if filtered.empty:
        return pd.DataFrame([columns])

    rank_map: dict[tuple[str, str], pd.Series] = {}
    for test in filtered["test"].unique():
        for metric in metrics:
            mask = (filtered["test"] == test) & (filtered["metric"] == metric)
            ranked = filtered.loc[mask, "p_value"].rank(method="min")
            rank_map[(test, metric)] = ranked

    rows = [columns]
    for test_name in ["t-test", "wilcoxon"]:
        test_rows = []
        for family in ["GPT-4o", "Llama3.1-70B", "Llama3.1-8B"]:
            base_model = FAMILY_COMPARISONS.get(family, {}).get("base")
            for comparison in comparisons:
                target_model = _model_for_comparison(family, comparison)
                if not base_model or not target_model:
                    continue
                row = [None] * len(columns)
                row[0] = base_model
                row[1] = target_model
                row[3] = "T-test" if test_name == "t-test" else "Signed Rank"
                row[4] = 0.05
                for metric, start in zip(metrics, [6, 11, 16, 21]):
                    metric_row = filtered[
                        (filtered["family"] == family)
                        & (filtered["comparison"] == comparison)
                        & (filtered["metric"] == metric)
                        & (filtered["test"] == test_name)
                    ]
                    if metric_row.empty:
                        continue
                    metric_row = metric_row.iloc[0]
                    row[start] = metric_row.get("p_value")
                    rank_series = rank_map.get((test_name, metric), pd.Series(dtype=float))
                    row[start + 1] = rank_series.get(metric_row.name, float("nan"))
                    adj_val = metric_row.get("adj_p")
                    row[start + 2] = adj_val
                    row[start + 3] = "yes" if pd.notna(adj_val) and float(adj_val) < 0.05 else "no"
                test_rows.append(row)
        # Order rows by accuracy p-value rank (Original Order) within test
        def _rank_val(row):
            vals = []
            for idx in [6, 11, 16, 21]:
                try:
                    vals.append(float(row[idx]))
                except Exception:
                    continue
            return min(vals) if vals else float("inf")
        test_rows.sort(key=_rank_val)
        for idx, row in enumerate(test_rows, start=1):
            row[2] = idx
        rows.extend(test_rows)
        rows.append([None] * len(columns))

    return pd.DataFrame(rows)


def _highlight_individquest_pvalues(path: Path) -> None:
    wb = load_workbook(path)
    if "IndividQuest_FisherExact(Tab3)" not in wb.sheetnames:
        return
    ws = wb["IndividQuest_FisherExact(Tab3)"]
    header = [cell.value for cell in ws[1]]
    p_cols = [idx + 1 for idx, val in enumerate(header) if val in {"Accuracy P", "Precision P", "Recall P"}]
    if not p_cols:
        wb.save(path)
        return
    fill_005 = PatternFill(start_color="CFE8FF", end_color="CFE8FF", fill_type="solid")
    fill_001 = PatternFill(start_color="7FC3FF", end_color="7FC3FF", fill_type="solid")
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for p_idx in p_cols:
            rank_idx = p_idx + 1
            adj_idx = p_idx + 2
            sig_idx = p_idx + 3
            adj_cell = row[adj_idx - 1]
            try:
                adj_val = float(adj_cell.value)
            except (TypeError, ValueError):
                adj_val = None
            if adj_val is None:
                continue
            if adj_val < 0.01:
                fill = fill_001
            elif adj_val < 0.05:
                fill = fill_005
            else:
                fill = None
            if fill:
                row[rank_idx - 1].fill = fill
                adj_cell.fill = fill
                row[sig_idx - 1].fill = fill
    # Apply percent formatting for percent-based columns
    percent_cols = [idx + 1 for idx, val in enumerate(header) if val in {"PcntTrue", "Accuracy", "Precision", "Recall", "F1"}]
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for col_idx in percent_cols:
            cell = row[col_idx - 1]
            if isinstance(cell.value, (int, float)):
                cell.number_format = "0.00%"
    if "ModelComp_SignedRankFig4" in wb.sheetnames:
        ws2 = wb["ModelComp_SignedRankFig4"]
        from openpyxl.styles import Font
        metric_titles = ["Accuracy", "Precision", "Recall", "F1"]
        title_row = 1
        title_starts = [2, 7, 12, 17]
        title_ends = [5, 10, 15, 20]
        for title, start, end in zip(metric_titles, title_starts, title_ends):
            ws2.merge_cells(start_row=title_row, start_column=start, end_row=title_row, end_column=end)
            cell = ws2.cell(row=title_row, column=start)
            cell.value = title
            cell.font = Font(size=18, bold=True)
        percent_cols_signed = [3, 4, 5, 8, 9, 10, 13, 14, 15, 18, 19, 20]
        skip_labels = {"P (T-Test)", "P (SignedRank)"}
        for row in ws2.iter_rows(min_row=1, max_row=ws2.max_row):
            labels = {row[idx - 1].value for idx in [2, 7, 12, 17]}
            if labels & skip_labels:
                continue
            for col_idx in percent_cols_signed:
                cell = row[col_idx - 1]
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "0.0%"
    wb.save(path)


def write_outputs(
    metrics: pd.DataFrame,
    details: List[dict],
    scenario_qid_frames: dict[str, pd.DataFrame],
) -> None:
    # Metrics: write CSV only (avoid duplicate Excel outputs)
    metrics.to_csv(config.OUTPUT_METRICS, index=False, encoding="utf-8-sig")

    # Detail rows workbook
    detail_df = pd.DataFrame(details)
    detail_df.sort_values(["sort_key"], inplace=True)
    detail_df.drop(columns=["sort_key"], inplace=True, errors="ignore")
    detail_sheets: dict[str, pd.DataFrame] = {"All": detail_df}
    _write_excel(config.DETAIL_METRICS_HUMAN.with_suffix(".xlsx"), detail_sheets)

    # Per-QID metrics workbook
    combined_qid = pd.concat(scenario_qid_frames.values(), ignore_index=True) if scenario_qid_frames else pd.DataFrame()
    if not combined_qid.empty:
        combined_qid.to_csv(config.OUTPUT_METRICS_BY_QID, index=False, encoding="utf-8-sig")


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--limit", type=int, default=None, help="Optional limit on number of rows.")
    parser.add_argument("--merged-path", type=Path, default=None, help="Override merged answers path.")
    parser.add_argument("--gpt5-path", type=Path, default=None, help="Override GPT-5 responses path.")
    parser.add_argument("--output-dir", type=Path, default=None, help="Directory for metrics/results outputs.")
    parser.add_argument("--figures-dir", type=Path, default=None, help="Directory for figure outputs.")
    parser.add_argument("--output-suffix", type=str, default="", help="Suffix appended to output filenames (e.g., new30).")
    args = parser.parse_args()

    suffix = args.output_suffix.strip()
    suffix = f"_{suffix}" if suffix else ""
    if args.merged_path:
        config.MERGED_PATH = args.merged_path
    if args.gpt5_path:
        config.GPT5_PATH = args.gpt5_path
    if args.output_dir:
        config.OUTPUT_DIR = args.output_dir
    config.OUTPUT_METRICS = config.OUTPUT_DIR / f"evaluation_metrics{suffix}.csv"
    config.OUTPUT_METRICS_BY_QID = config.OUTPUT_DIR / f"evaluation_metrics_by_qid{suffix}.csv"
    config.STAT_RESULTS = config.OUTPUT_DIR / f"statistical_tests{suffix}.xlsx"
    config.DETAIL_METRICS_HUMAN = config.OUTPUT_DIR / f"detailed_evaluation{suffix}.csv"
    if args.figures_dir:
        config.OUTPUT_TABLE_DIR = args.figures_dir

    logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
    (
        metrics,
        details,
        figures,
        qid_rows,
        scenario_qid_frames,
    ) = run(args.limit)
    if metrics.empty:
        logging.error("No scenarios produced metrics.")
        return 1
    write_outputs(metrics, details, scenario_qid_frames)

    qid_df = pd.DataFrame(qid_rows)
    if not qid_df.empty:
        # Persist the exact per-QID precision/recall data used for plotting
        precision_recall_csv = config.OUTPUT_DIR / f"precision_recall_by_qid{suffix}.csv"
        qid_df.to_csv(precision_recall_csv, index=False, encoding="utf-8-sig")
        # Re-read from disk so plots are guaranteed to use the saved artifact
        qid_df = pd.read_csv(precision_recall_csv)

    overall_stats = {}
    fisher_df = pd.DataFrame()
    pair_df = pd.DataFrame()
    overall_qid_df = next(iter(scenario_qid_frames.values()), pd.DataFrame())
    fisher_metrics = ["accuracy", "precision", "recall", "f1"]
    sig_map: dict[str, dict[tuple[str, str], float]] = {}
    if overall_qid_df is not None and not overall_qid_df.empty:
        paired_metrics = ["accuracy", "precision", "recall", "f1"]
        fisher_df = stat_utils.compute_fisher_tests(
            overall_qid_df,
            FAMILY_COMPARISONS,
            fisher_metrics,
        )
        pair_df, overall_stats_raw, _ = stat_utils.compute_pairwise_tests(
            overall_qid_df,
            FAMILY_COMPARISONS,
            paired_metrics,
        )
        # Build significance map from adjusted Wilcoxon p-values (<0.05) for plotting
        if not pair_df.empty:
            w = pair_df[pair_df["test"] == "wilcoxon"]
            for _, row in w.iterrows():
                metric = row.get("metric")
                family = row.get("family")
                comparison = row.get("comparison")
                adj_p = row.get("adj_p")
                if metric is None or family is None or comparison is None or adj_p is None:
                    continue
                try:
                    val = float(adj_p)
                except Exception:
                    continue
                if val < 0.05:
                    sig_map.setdefault(str(metric), {})[(str(family), str(comparison))] = val
        overall_stats = sig_map
    logging.info("Wrote metrics to %s", config.OUTPUT_METRICS)
    logging.info("Wrote detail rows to %s", config.DETAIL_METRICS_HUMAN)
    if not pair_df.empty or not fisher_df.empty:
        fisher_qid_sheet = _build_fisher_qid_sheet(fisher_df)
        table3_df = _build_table3(fisher_qid_sheet, overall_qid_df)
        with pd.ExcelWriter(config.STAT_RESULTS, engine="openpyxl") as writer:
            if not pair_df.empty:
                _drop_qid_cols(pair_df).to_excel(writer, sheet_name="Paired Tests", index=False)
            if not fisher_df.empty:
                qid_sheet = fisher_qid_sheet
                if not qid_sheet.empty:
                    qid_sheet.to_excel(writer, sheet_name="Fisher Exact Test", index=False)
            if table3_df is not None and not table3_df.empty:
                table3_df.to_excel(writer, sheet_name="Table3", index=False)
        logging.info("Wrote combined statistical tests to %s", config.STAT_RESULTS)
        if overall_qid_df is not None and not overall_qid_df.empty:
            supp_path = config.OUTPUT_DIR / "SuppFile2_Stats_generated.xlsx"
            individ_df = _build_suppfile2_individquest(
                overall_qid_df,
                fisher_qid_sheet,
                comparisons=["FT", "QSP"],
            )
            signedrank_df = _build_suppfile2_signedrank(overall_qid_df, pair_df)
            bh_df = _build_suppfile2_modelcomp_bh(pair_df)
            with pd.ExcelWriter(supp_path, engine="openpyxl") as writer:
                individ_df.to_excel(writer, sheet_name="IndividQuest_FisherExact(Tab3)", index=False)
                signedrank_df.to_excel(writer, sheet_name="ModelComp_SignedRankFig4", index=False, header=False)
                bh_df.to_excel(writer, sheet_name="ModelComp_BH_AdjustFig4", index=False, header=False)
            _highlight_individquest_pvalues(supp_path)
            logging.info("Wrote SuppFile2 stats to %s", supp_path)
    # Export aggregated Fisher results alongside metrics
    fisher_summary = _aggregate_fisher_summary(
        overall_qid_df if overall_qid_df is not None else pd.DataFrame(),
        FAMILY_COMPARISONS,
        fisher_metrics,
    )
    scenario_metrics = metrics.copy()
    fisher_col_map = {
        "accuracy": ("p_acc_fisher", "adj_p_acc_fisher"),
        "precision": ("p_prec_fisher", "adj_p_prec_fisher"),
        "recall": ("p_rec_fisher", "adj_p_rec_fisher"),
        "f1": ("p_f1_fisher", "adj_p_f1_fisher"),
    }
    desired_order = [
        "samples",
        "model",
        "tp",
        "tn",
        "fp",
        "fn",
        "accuracy",
        "p_acc_fisher",
        "adj_p_acc_fisher",
        "precision",
        "p_prec_fisher",
        "adj_p_prec_fisher",
        "recall",
        "p_rec_fisher",
        "adj_p_rec_fisher",
        "f1",
        "p_f1_fisher",
        "adj_p_f1_fisher",
    ]
    if not scenario_metrics.empty:
        for metric_name, (p_col, adj_col) in fisher_col_map.items():
            scenario_metrics[p_col] = np.nan
            scenario_metrics[adj_col] = np.nan
    if not scenario_metrics.empty and not fisher_summary.empty:
        for _, row in fisher_summary.iterrows():
            target_model = row["target_model"]
            metric_name = row["metric"]
            p_val = row.get("p_value")
            adj_val = row.get("adj_p_value")
            mask = scenario_metrics["model"] == target_model
            if mask.any():
                p_col, adj_col = fisher_col_map.get(metric_name, (None, None))
                if p_col:
                    scenario_metrics.loc[mask, p_col] = p_val
                if adj_col:
                    scenario_metrics.loc[mask, adj_col] = adj_val
    if not scenario_metrics.empty:
        for col in desired_order:
            if col not in scenario_metrics.columns:
                scenario_metrics[col] = np.nan
        scenario_metrics = scenario_metrics.reindex(columns=desired_order)
        fisher_metrics_path = config.OUTPUT_DIR / f"evaluation_metrics_fisher{suffix}.xlsx"
        scenario_metrics.to_excel(fisher_metrics_path, index=False)
        logging.info("Wrote metrics with Fisher p-values to %s", fisher_metrics_path)
    dataset_label = _dataset_label_from_suffix(args.output_suffix)
    details_path = config.OUTPUT_DIR / f"detailed_evaluation{suffix}.xlsx"
    for display_title, scenario_title, subset, scenario_qid_df in figures:
        sig = overall_stats
        if suffix:
            base_name = suffix.strip("_") or "figure"
        else:
            base_name = slugify(dataset_label)
        full_title = dataset_label or scenario_title
        generate_figures(
            subset,
            scenario_title,
            config.OUTPUT_TABLE_DIR,
            significance=sig,
            comparisons=FAMILY_COMPARISONS,
            base_name=base_name,
            display_title=full_title,
        )
        ci_df = build_bar_chart_confidence_intervals(details_path, subset)
        ci_path = config.OUTPUT_TABLE_DIR / f"{base_name}-bar-chart-confidence-intervals.csv"
        ci_df.to_csv(ci_path, index=False, encoding="utf-8-sig")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
