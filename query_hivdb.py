import mysql.connector
from mysql.connector import Error
import re
import os
from pathlib import Path
from dotenv import load_dotenv
from dotenv import find_dotenv
load_dotenv(find_dotenv())
import openpyxl
from operator import itemgetter
import csv
from collections import defaultdict


WS = Path(__file__).resolve().parent
SQL_PATH = WS / 'sql'
PAPER_FILE = WS / 'database' / 'HIVDB_tblReferences.xlsx'
# PAPER_SAVE_FILE = WS / 'database' / 'HIVDB_tblReferences_answers.xlsx'
PAPER_SAVE_FILE = WS / 'database' / 'HIVDB_tblReferences_answers.csv'


def dump_csv(file_path, table, headers=[], remain=True):

    file_path = Path(file_path)

    table_headers = []
    for rec in table:
        for key in rec.keys():
            if key not in table_headers:
                table_headers.append(key)

    if not headers:
        headers = table_headers
    else:
        remain_headers = [
            i
            for i in table_headers
            if i not in headers
        ]
        if remain:
            headers = headers + remain_headers
        table = [
            {
                k: v
                for k, v in i.items()
                if k in headers
            }
            for i in table
        ]

    file_path.parent.mkdir(exist_ok=True, parents=True)

    with open(file_path, 'w', encoding='utf-8-sig') as fd:
        writer = csv.DictWriter(fd, fieldnames=headers)
        writer.writeheader()
        writer.writerows(table)


def create_db_connection(host_name, user_name, user_password, db_name, port):
    connection = None
    try:
        connection = mysql.connector.connect(
            host=host_name,
            user=user_name,
            password=user_password,
            database=db_name,
            port=port
        )
        print("MySQL Database connection successful")
    except Error as err:
        print(f"Error: '{err}'")
    return connection


def read_query(connection, query):
    cursor = connection.cursor()
    result = None
    try:
        cursor.execute(query)
        result = cursor.fetchall()
        return result
    except Error as err:
        print(f"Error: '{err}'")


host = os.getenv('DB_HOST')
user = os.getenv('DB_USER')
password = os.getenv('DB_PASSWORD')
database = os.getenv('DB_DATABASE')
port = os.getenv('DB_PORT')


def load_paper():
    workbook = openpyxl.load_workbook(str(PAPER_FILE))
    sheet = workbook.active

    table = []
    header = []
    for idx, row in enumerate(sheet.iter_rows(values_only=True)):
        if idx == 0:
            header = row
            continue

        table.append(dict(zip(header, row)))

    return table


def dump_paper(table):
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    header = list(table[0].keys())

    sheet.append(header)

    table = [
        [r.get(h, '') for h in header]
        for r in table
    ]

    for row in table:
        sheet.append(row)

    workbook.save(str(PAPER_SAVE_FILE))


def load_sql():
    sql_list = defaultdict(dict)
    for i in SQL_PATH.iterdir():
        if i.suffix != '.sql':
            continue

        qid = i.stem.replace('b', '')
        with open(i) as fd:

            if i.stem.endswith('b'):
                sql_list[qid]['sql_b'] = fd.read().strip()
            else:
                sql_list[qid]['sql'] = fd.read().strip()

    sql_list = [
        {
            'QID': k,
            'sql': v['sql'],
            'sql_b': v.get('sql_b', v['sql']),
        }
        for k, v in sql_list.items()
    ]

    sql_list.sort(key=lambda x: int(x['QID'][1:]))

    return sql_list


def work():
    db = create_db_connection(host, user, password, database, port)

    papers = load_paper()

    sql_list = load_sql()

    for idx, p in enumerate(papers):

        # Q0 is for switch main tables
        sql = sql_list[0]['sql'].format(pubmed_id=p['MedlineID'])

        result = read_query(db, sql)

        if result[0][0] > 0:
            key = 'sql'
        else:
            key = 'sql_b'

        for s in sql_list[1:]:
            qid = s['QID']
            sql = s[key]
            sql = sql.format(pubmed_id=p['MedlineID'])
            result = read_query(db, sql)

            if len(result) > 10:
                result = result[:3] + [[f"({len(result)} results)"]]

            answer = ';'.join([
                '\n'.join([str(v) for v in r])
                for r in result
            ])

            if int(qid[1:]) in [10, 11, 12]:
                answer = answer.replace(
                        'MC', 'molecular clone'
                    ).replace(
                        'BC', 'SGS'
                    ).replace(
                        'Unknown', 'Cloned'
                    )

            if int(qid[1:]) in [9]:
                answer = answer.replace(
                        'Dideoxy', 'Sanger'
                    )

            p[f'{qid} Ans'] = answer

        if idx % 10 == 0 and idx > 0:
            print(idx)

        # if idx > 10:
        #     break

    dump_csv(PAPER_SAVE_FILE, papers)


if __name__ == '__main__':
    work()
