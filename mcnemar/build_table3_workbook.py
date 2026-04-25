from __future__ import annotations

import argparse
import sys
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.append(str(ROOT))

from eval.evaluation import FAMILY_COMPARISONS, _build_fisher_qid_sheet, _model_for_comparison  # type: ignore
from mcnemar.build_suppfile2_workbook import DEFAULT_COMPARISONS, _compute_mcnemar_tests  # type: ignore


DEFAULT_SUFFIX = "full150"
DEFAULT_OUTPUT = Path(__file__).resolve().parent / "Table3_formatted.xlsx"
FAMILY_ORDER = ["GPT-4o", "Llama3.1-70B", "Llama3.1-8B"]


def _metric_lookup(qid_df: pd.DataFrame) -> dict[tuple[int, str], dict[str, float]]:
    lookup: dict[tuple[int, str], dict[str, float]] = {}
    for _, row in qid_df.iterrows():
        qid = int(row["QID"])
        model = str(row["model"])
        lookup[(qid, model)] = {
            "accuracy": float(row.get("accuracy", 0)),
            "precision": float(row.get("precision", 0)),
            "recall": float(row.get("recall", 0)),
        }
    return lookup


def _significance_map(long_df: pd.DataFrame, metrics: list[str]) -> tuple[dict[tuple[str, str, int], set[str]], dict[tuple[str, str, str, int], float]]:
    sig_map: dict[tuple[str, str, int], set[str]] = {}
    adj_map: dict[tuple[str, str, str, int], float] = {}
    for _, row in long_df.iterrows():
        family = str(row.get("family"))
        comparison = str(row.get("comparison"))
        metric = str(row.get("metric"))
        if comparison not in DEFAULT_COMPARISONS or metric not in metrics:
            continue
        try:
            adj_p = float(row.get("adj_p"))
            base = float(row.get("base"))
            target = float(row.get("target"))
            qid = int(row.get("QID"))
        except (TypeError, ValueError):
            continue
        if adj_p < 0.05 and target > base:
            sig_map.setdefault((family, comparison, qid), set()).add(metric)
            adj_map[(family, comparison, metric, qid)] = adj_p
    return sig_map, adj_map


def _format_metric_value(value: float | None, adj_p: float | None, significant: bool) -> str:
    if value is None:
        return ""
    suffix = ""
    if significant and adj_p is not None:
        if adj_p < 0.01:
            suffix = "**"
        elif adj_p < 0.05:
            suffix = "*"
    return f"{value * 100:.1f}{suffix}"


def _build_records(
    qid_df: pd.DataFrame,
    long_df: pd.DataFrame,
    metrics: list[str],
) -> dict[str, list[dict[str, object]]]:
    qid_question = qid_df.groupby("QID")["Question"].first().to_dict()
    metric_lookup = _metric_lookup(qid_df)
    sig_map, adj_map = _significance_map(long_df, metrics)
    records_by_family: dict[str, list[dict[str, object]]] = {}

    for family in FAMILY_ORDER:
        base_model = FAMILY_COMPARISONS[family]["base"]
        ft_model = _model_for_comparison(family, "FT")
        qsp_model = _model_for_comparison(family, "QSP")
        sig_qids = sorted({qid for fam, _, qid in sig_map if fam == family})
        if not sig_qids:
            records_by_family[family] = []
            continue
        family_rows: list[dict[str, object]] = []
        for qid in sig_qids:
            row: dict[str, object] = {
                "QID": qid,
                "Question": qid_question.get(qid, ""),
            }
            base_vals = metric_lookup.get((qid, base_model), {})
            for metric in metrics:
                row[f"base_{metric}"] = _format_metric_value(base_vals.get(metric), None, False)
            for label, model_name in [("FT", ft_model), ("QSP", qsp_model)]:
                vals = metric_lookup.get((qid, model_name), {}) if model_name else {}
                sig_metrics = sig_map.get((family, label, qid), set())
                for metric in metrics:
                    row[f"{label}_{metric}"] = _format_metric_value(
                        vals.get(metric),
                        adj_map.get((family, label, metric, qid)),
                        metric in sig_metrics,
                    )
            family_rows.append(row)
        records_by_family[family] = family_rows
    return records_by_family


def _write_sheet(ws, title: str, metrics: list[str], records_by_family: dict[str, list[dict[str, object]]]) -> None:
    metric_titles = {"accuracy": "Accuracy", "precision": "Precision", "recall": "Recall"}
    col_count = 2 + 3 * len(metrics)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_count)
    ws.cell(1, 1).value = title
    ws.cell(1, 1).font = Font(bold=True)
    ws.cell(1, 1).alignment = Alignment(wrap_text=True)

    current_col = 3
    for metric in metrics:
        ws.merge_cells(start_row=2, start_column=current_col, end_row=2, end_column=current_col + 2)
        ws.cell(2, current_col).value = metric_titles[metric]
        ws.cell(2, current_col).font = Font(bold=True)
        current_col += 3

    current_col = 3
    for _ in metrics:
        for label in ["B", "FT", "QSP"]:
            ws.cell(3, current_col).value = label
            ws.cell(3, current_col).font = Font(bold=True)
            current_col += 1

    row_idx = 4
    for family in FAMILY_ORDER:
        family_rows = records_by_family.get(family, [])
        if not family_rows:
            continue
        ws.cell(row_idx, 1).value = family
        ws.cell(row_idx, 1).font = Font(bold=True)
        row_idx += 1
        for record in family_rows:
            ws.cell(row_idx, 1).value = record["QID"]
            ws.cell(row_idx, 2).value = record["Question"]
            current_col = 3
            for metric in metrics:
                ws.cell(row_idx, current_col).value = record[f"base_{metric}"]
                ws.cell(row_idx, current_col + 1).value = record[f"FT_{metric}"]
                ws.cell(row_idx, current_col + 2).value = record[f"QSP_{metric}"]
                current_col += 3
            row_idx += 1

    ws.cell(row_idx, 1).value = "Footnote: **: adjusted p < 0.01; *: adjusted p < 0.05"
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 90
    for col in range(3, col_count + 1):
        ws.column_dimensions[chr(64 + col)].width = 10


def build_workbook(output_path: Path, suffix: str) -> Path:
    qid_df = pd.read_csv(ROOT / "eval" / "results" / f"evaluation_metrics_by_qid_{suffix}.csv")
    fisher_long = pd.read_excel(ROOT / "eval" / "results" / f"statistical_tests_{suffix}.xlsx", sheet_name="Fisher Exact Test")
    detail_df = pd.read_excel(ROOT / "eval" / "results" / f"detailed_evaluation_{suffix}.xlsx", sheet_name="All")
    mcnemar_long = _build_fisher_qid_sheet(
        _compute_mcnemar_tests(
            qid_df,
            detail_df,
            FAMILY_COMPARISONS,
            ["accuracy", "recall"],
            bh_comparisons=DEFAULT_COMPARISONS,
        )
    )

    fisher_records = _build_records(qid_df, fisher_long, ["accuracy", "precision", "recall"])
    mcnemar_records = _build_records(qid_df, mcnemar_long, ["accuracy", "recall"])

    wb = Workbook()
    ws_fisher = wb.active
    ws_fisher.title = "FisherExactTest"
    _write_sheet(
        ws_fisher,
        "Table 3. Accuracy, Precision, and Recall for the Research Questions for which an Improvement was Observed After Fine-Tuning (FT) or Question-Specific Prompting (QSP) using Fisher Exact Test",
        ["accuracy", "precision", "recall"],
        fisher_records,
    )

    ws_mcnemar = wb.create_sheet("McNemarTest")
    _write_sheet(
        ws_mcnemar,
        "Table 3. Accuracy and Recall for the Research Questions for which an Improvement was Observed After Fine-Tuning (FT) or Question-Specific Prompting (QSP) using McNemar Test",
        ["accuracy", "recall"],
        mcnemar_records,
    )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--suffix", default=DEFAULT_SUFFIX, help="Dataset suffix (default: full150).")
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT, help="Output workbook path.")
    args = parser.parse_args()
    build_workbook(args.output, args.suffix)
    print(args.output)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
