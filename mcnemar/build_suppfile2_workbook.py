from __future__ import annotations

import argparse
import sys
from pathlib import Path
from typing import Iterable

import pandas as pd
from openpyxl import load_workbook
from scipy.stats import binomtest

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.append(str(ROOT))

from eval import statistics as stat_utils  # type: ignore
from eval.evaluation import (  # type: ignore
    FAMILY_COMPARISONS,
    _build_fisher_qid_sheet,
    _build_suppfile2_individquest,
    _build_suppfile2_modelcomp_bh,
    _build_suppfile2_signedrank,
    _highlight_individquest_pvalues,
    _model_for_comparison,
)
from eval.normalize import canonicalize_answer, is_empty_token  # type: ignore


DEFAULT_SUFFIX = "full150"
DEFAULT_COMPARISONS = ("FT", "QSP")


def _round_sig(val: float) -> float:
    try:
        return float(f"{float(val):.3g}")
    except Exception:
        return float(val)


def _exact_mcnemar_p(base_values: pd.Series, target_values: pd.Series) -> float:
    base = base_values.astype(int)
    target = target_values.astype(int)
    target_only = int(((base == 0) & (target == 1)).sum())
    base_only = int(((base == 1) & (target == 0)).sum())
    discordant = target_only + base_only
    if discordant == 0:
        return 1.0
    return float(binomtest(min(target_only, base_only), n=discordant, p=0.5).pvalue)


def _reference_positive(row: pd.Series) -> bool:
    qtype = str(row.get("Type", "")).strip().lower()
    ref_norm = canonicalize_answer(str(row.get("Human Answer", "")))
    if qtype == "boolean":
        return ref_norm == "yes"
    if qtype == "number":
        return not is_empty_token(ref_norm, allow_zero=False)
    return not is_empty_token(ref_norm)


def _metric_lookup(qid_df: pd.DataFrame) -> dict[tuple[int, str], dict[str, float]]:
    lookup: dict[tuple[int, str], dict[str, float]] = {}
    for _, row in qid_df.iterrows():
        qid = int(row["QID"])
        model = str(row["model"])
        lookup[(qid, model)] = {
            "tp": float(row.get("tp", 0)),
            "fn": float(row.get("fn", 0)),
            "fp": float(row.get("fp", 0)),
            "tn": float(row.get("tn", 0)),
            "accuracy": float(row.get("accuracy", 0)),
            "precision": float(row.get("precision", 0)),
            "recall": float(row.get("recall", 0)),
            "f1": float(row.get("f1", 0)),
            "samples": float(row.get("samples", 0)),
        }
    return lookup


def _compute_mcnemar_tests(
    qid_df: pd.DataFrame,
    detail_df: pd.DataFrame,
    comparisons: dict,
    metrics: Iterable[str],
) -> pd.DataFrame:
    if qid_df.empty or detail_df.empty:
        return pd.DataFrame()

    detail = detail_df.copy()
    detail["QID"] = detail["QID"].astype(int)
    detail["__ref_positive"] = detail.apply(_reference_positive, axis=1)
    qid_lookup = _metric_lookup(qid_df)
    all_qids = sorted(qid_df["QID"].astype(int).unique().tolist())
    records: list[dict] = []

    for family, mapping in comparisons.items():
        base_model = mapping["base"]
        base_correct_col = f"{base_model} Correct"
        if base_correct_col not in detail.columns:
            continue
        for target_model in mapping["targets"]:
            target_correct_col = f"{target_model} Correct"
            if target_correct_col not in detail.columns:
                continue
            for metric in metrics:
                base_values: dict[int, float] = {}
                target_values: dict[int, float] = {}
                p_values: dict[int, float] = {}
                for qid in all_qids:
                    subset = detail[detail["QID"] == qid]
                    if subset.empty:
                        continue
                    if metric == "recall":
                        subset = subset[subset["__ref_positive"]]
                        if subset.empty:
                            continue
                    elif metric != "accuracy":
                        continue

                    base_metric = qid_lookup.get((qid, base_model), {}).get(metric)
                    target_metric = qid_lookup.get((qid, target_model), {}).get(metric)
                    if base_metric is None or target_metric is None:
                        continue

                    p_value = _exact_mcnemar_p(subset[base_correct_col], subset[target_correct_col])
                    base_values[qid] = _round_sig(base_metric)
                    target_values[qid] = _round_sig(target_metric)
                    p_values[qid] = _round_sig(p_value)

                if not p_values:
                    continue

                record = {
                    "family": family,
                    "comparison": target_model.replace(f"{family} ", ""),
                    "metric": metric,
                    "test": "mcnemar",
                }
                for qid in all_qids:
                    record[f"base_qid_{qid}"] = _round_sig(base_values.get(qid, float("nan")))
                    record[f"target_qid_{qid}"] = _round_sig(target_values.get(qid, float("nan")))
                    record[f"p_value_qid_{qid}"] = _round_sig(p_values.get(qid, float("nan")))
                    record[f"adj_p_qid_{qid}"] = float("nan")
                records.append(record)

    df = pd.DataFrame(records)
    if df.empty:
        return df

    for metric in metrics:
        metric_mask = df["metric"] == metric
        for qid in all_qids:
            p_col = f"p_value_qid_{qid}"
            adj_col = f"adj_p_qid_{qid}"
            if p_col not in df.columns:
                continue
            vals = df.loc[metric_mask, p_col]
            valid_mask = metric_mask & vals.notna()
            pvals = df.loc[valid_mask, p_col].astype(float).tolist()
            if not pvals:
                continue
            adjusted = stat_utils.benjamini_hochberg(pvals)
            df.loc[valid_mask, adj_col] = [_round_sig(val) for val in adjusted]

    for col in df.columns:
        if df[col].dtype.kind in {"f", "i"}:
            df[col] = df[col].apply(_round_sig)
    return df


def _build_suppfile2_individquest_mcnemar(
    qid_df: pd.DataFrame,
    mcnemar_qid_df: pd.DataFrame,
    comparisons: Iterable[str],
) -> pd.DataFrame:
    header = [
        "Question",
        "PcntTrue",
        "Model",
        "State",
        "TP",
        "FN",
        "FP",
        "TN",
        "Accuracy",
        "Precision",
        "Recall",
        "F1",
        "",
        "State",
        "TP",
        "FN",
        "FP",
        "TN",
        "Accuracy",
        "Precision",
        "Recall",
        "F1",
        "",
        "Accuracy P",
        "Rank",
        "Adjusted P",
        "BH sig",
        "",
        "Recall P",
        "Rank",
        "Adjusted P",
        "BH sig",
        "",
        "Base (base model)",
    ]
    if qid_df is None or qid_df.empty:
        return pd.DataFrame(columns=header)

    qid_counts = qid_df.groupby("QID").first()
    pcnt_true = (qid_counts["tp"] + qid_counts["fn"]) / qid_counts["samples"]
    metric_map = _metric_lookup(qid_df)

    rows: list[list[object]] = []
    qids = sorted(qid_df["QID"].astype(int).unique())
    questions_per_block = 16

    for family in ["GPT-4o", "Llama3.1-70B", "Llama3.1-8B"]:
        base_model = FAMILY_COMPARISONS.get(family, {}).get("base")
        if not base_model:
            continue
        for comparison in comparisons:
            target_model = _model_for_comparison(family, comparison)
            if not target_model:
                continue
            comp_df = mcnemar_qid_df[
                (mcnemar_qid_df["family"] == family)
                & (mcnemar_qid_df["comparison"] == comparison)
                & (mcnemar_qid_df["metric"].isin(["accuracy", "recall"]))
            ].copy()
            if comp_df.empty:
                continue
            comp_df["rank"] = comp_df.groupby("metric")["p_value"].rank(method="min")
            comp_df["bh_sig"] = comp_df["adj_p"].apply(lambda v: "yes" if pd.notna(v) and float(v) < 0.05 else "no")
            question_count = 0
            for qid in qids:
                base = metric_map.get((qid, base_model))
                target = metric_map.get((qid, target_model))
                if base is None or target is None:
                    continue

                row_vals: list[object] = [None] * len(header)
                row_vals[0] = qid
                row_vals[1] = round(float(pcnt_true.get(qid, float("nan"))), 4)
                row_vals[2] = family
                row_vals[3] = "Base"
                row_vals[4:12] = [
                    base["tp"],
                    base["fn"],
                    base["fp"],
                    base["tn"],
                    round(base["accuracy"], 4),
                    round(base["precision"], 4),
                    round(base["recall"], 4),
                    round(base["f1"], 4),
                ]
                row_vals[13] = comparison
                row_vals[14:22] = [
                    target["tp"],
                    target["fn"],
                    target["fp"],
                    target["tn"],
                    round(target["accuracy"], 4),
                    round(target["precision"], 4),
                    round(target["recall"], 4),
                    round(target["f1"], 4),
                ]

                for metric, offset in [("accuracy", 23), ("recall", 28)]:
                    metric_row = comp_df[(comp_df["QID"] == qid) & (comp_df["metric"] == metric)]
                    if metric_row.empty:
                        continue
                    metric_row = metric_row.iloc[0]
                    row_vals[offset] = metric_row.get("p_value")
                    row_vals[offset + 1] = metric_row.get("rank")
                    row_vals[offset + 2] = metric_row.get("adj_p")
                    row_vals[offset + 3] = metric_row.get("bh_sig")
                rows.append(row_vals)
                question_count += 1
                if question_count % questions_per_block == 0:
                    rows.append([None] * len(header))
            if question_count % questions_per_block != 0:
                rows.append([None] * len(header))
        rows.append([None] * len(header))

    legend = {
        0: "Base (base model)",
        1: "FT (fine-tuned model)",
        3: "P (p value)",
        4: "BH sig (Benjamini Hochberg): Is adjusted p value <0.05?",
        5: "Cyan: unadjusted p value <0.01; Light blue: unadjusted p value <0.05",
        8: "TP (true positive)",
        9: "TN (true negative)",
        10: "FP (false positive)",
        11: "FN (false negative)",
    }
    for idx, text in legend.items():
        if idx < len(rows):
            rows[idx][-1] = text
    return pd.DataFrame(rows, columns=header)


def _highlight_sheet(path: Path, sheet_name: str) -> None:
    wb = load_workbook(path)
    if sheet_name not in wb.sheetnames:
        wb.save(path)
        return
    ws = wb[sheet_name]
    header = [cell.value for cell in ws[1]]
    p_cols = [
        idx + 1
        for idx, val in enumerate(header)
        if val in {"Accuracy P", "Precision P", "Recall P"}
    ]
    fill_005 = "CFE8FF"
    fill_001 = "7FC3FF"
    from openpyxl.styles import PatternFill

    low_fill = PatternFill(start_color=fill_005, end_color=fill_005, fill_type="solid")
    high_fill = PatternFill(start_color=fill_001, end_color=fill_001, fill_type="solid")

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for p_idx in p_cols:
            rank_idx = p_idx + 1
            adj_idx = p_idx + 2
            sig_idx = p_idx + 3
            adj_cell = row[adj_idx - 1]
            try:
                adj_val = float(adj_cell.value)
            except (TypeError, ValueError):
                continue
            if adj_val < 0.01:
                fill = high_fill
            elif adj_val < 0.05:
                fill = low_fill
            else:
                fill = None
            if fill is not None:
                row[rank_idx - 1].fill = fill
                adj_cell.fill = fill
                row[sig_idx - 1].fill = fill

    percent_cols = [
        idx + 1
        for idx, val in enumerate(header)
        if val in {"PcntTrue", "Accuracy", "Precision", "Recall", "F1"}
    ]
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for col_idx in percent_cols:
            cell = row[col_idx - 1]
            if isinstance(cell.value, (int, float)):
                cell.number_format = "0.00%"
    wb.save(path)


def build_workbook(output_path: Path, suffix: str) -> Path:
    results_dir = ROOT / "eval" / "results"
    qid_path = results_dir / f"evaluation_metrics_by_qid_{suffix}.csv"
    detail_path = results_dir / f"detailed_evaluation_{suffix}.xlsx"

    qid_df = pd.read_csv(qid_path)
    detail_df = pd.read_excel(detail_path, sheet_name="All")

    fisher_df = stat_utils.compute_fisher_tests(
        qid_df,
        FAMILY_COMPARISONS,
        ["accuracy", "precision", "recall", "f1"],
    )
    fisher_qid_df = _build_fisher_qid_sheet(fisher_df)
    pair_df, _, _ = stat_utils.compute_pairwise_tests(
        qid_df,
        FAMILY_COMPARISONS,
        ["accuracy", "precision", "recall", "f1"],
    )
    mcnemar_df = _compute_mcnemar_tests(
        qid_df,
        detail_df,
        FAMILY_COMPARISONS,
        ["accuracy", "recall"],
    )
    mcnemar_qid_df = _build_fisher_qid_sheet(mcnemar_df)

    mcnemar_sheet = _build_suppfile2_individquest_mcnemar(
        qid_df,
        mcnemar_qid_df,
        comparisons=DEFAULT_COMPARISONS,
    )
    fisher_sheet = _build_suppfile2_individquest(qid_df, fisher_qid_df, comparisons=DEFAULT_COMPARISONS)
    signedrank_sheet = _build_suppfile2_signedrank(qid_df, pair_df)
    bh_sheet = _build_suppfile2_modelcomp_bh(pair_df)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        fisher_sheet.to_excel(writer, sheet_name="IndividQuest_FisherExact(Tab3)", index=False)
        mcnemar_sheet.to_excel(writer, sheet_name="IndividQuest_McNemar(Tab3)", index=False)
        signedrank_sheet.to_excel(writer, sheet_name="ModelComp_SignedRankFig4", index=False, header=False)
        bh_sheet.to_excel(writer, sheet_name="ModelComp_BH_AdjustFig4", index=False, header=False)

    _highlight_individquest_pvalues(output_path)
    _highlight_sheet(output_path, "IndividQuest_McNemar(Tab3)")
    return output_path


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--suffix",
        default=DEFAULT_SUFFIX,
        help="Dataset suffix matching eval/results artifacts (default: full150).",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=Path(__file__).resolve().parent / "SuppFile2_Stats_generated.xlsx",
        help="Output workbook path.",
    )
    args = parser.parse_args()

    build_workbook(args.output, args.suffix)
    print(args.output)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
