import numpy as np
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from math import sqrt


def load_excel(excel_file_path, main_sheet=None):

    workbook = load_workbook(str(excel_file_path), read_only=True)

    sheet = None
    for sheet_name in workbook.sheetnames:
        if main_sheet:
            if sheet_name != main_sheet:
                continue
            sheet = workbook[sheet_name]
            break
        else:
            sheet = workbook[sheet_name]
            break

    sheet = workbook[sheet_name]
    table = []

    header = []
    for idx, i in enumerate(sheet):
        if idx == 0:
            for j in i:
                header.append(j.value)
            continue

        row = []
        for j in i:
            value = j.value
            if j.number_format and j.number_format.endswith('%'):
                value = f"{value * 100}%"
            if value is None:
                value = ''
            row.append(value)

        # Skip blank row
        if not any(row):
            continue
        table.append(dict(zip(header, row)))

    return table


def draw_baseline(file_name, table, bar_color, **options):

    fig, ax = plt.subplots(figsize=(15, 8))
    ax.set_ylim(50, options['max_tick'] + 23)
    ax.set_yticks(range(50, 110, 10))
    ax.set_xlim(-0.2, 3.8)
    ax.set_ylabel('Percentage', fontsize=14)
    ax.tick_params(axis='both', labelsize=12)

    width = 0.27
    gap = 0.02
    xticks = []
    xlabels = []
    group_labels = []

    for i, group_name in enumerate(['Accuracy', 'Precision', 'Recall', 'F1']):
        row = table[i]

        for j, label in enumerate(['GPT-4', 'Llama 3.1 70B', 'Llama 3.1 8B']):
            x = i + j * width + j * gap
            ax.bar(
                x, row[label], width, color=bar_color)
            ax.text(
                x, row[label] * 1.02, f'{row[label]:.0f}%', ha='center', va='bottom')
            xticks.append(x)
            xlabels.append(label)

        group_labels.append(
            (
                i + width + gap,
                group_name
            )
        )

        text_y = options['max_tick'] + 5

        # GPT-4 vs 70B
        if is_significant(row['S1']):
            ax.plot(
                [i,         i,              i + width,          i + width],
                [text_y,    text_y * 1.03,  text_y * 1.03,      text_y],
                color='black',
                lw=1.5)
            ax.text(
                i + width / 2 + gap / 2, text_y * 1.035,
                # f"mean={row['M1']}\nstd={row['STD1']}\np{'' if '<' in str(row['S1']) else '='}{row['S1']}",
                f"p{'' if '<' in str(row['S1']) else '='}{row['S1']}",
                ha='center', va='bottom')

        if is_significant(row['S3']):
            # 70B vs 8B
            ax.plot(
                [i + width + 2 * gap,   i + width + 2 * gap,    i + 2 * width + 2 * gap,    i + 2 * width + 2 * gap],
                [text_y,                text_y * 1.03,          text_y * 1.03,              text_y],
                color='black',
                lw=1.5)
            ax.text(
                i + width + width / 2 + gap + gap / 2, text_y * 1.035,
                # f"mean={row['M3']}\nstd={row['STD3']}\np{'' if '<' in str(row['S3']) else '='}{row['S3']}",
                f"p{'' if '<' in str(row['S3']) else '='}{row['S3']}",
                ha='center', va='bottom')

        text_y = text_y + 7
        if is_significant(row['S2']):
            # GPT-4 vs 8B
            ax.plot(
                [i,         i,              i + 2 * width + 2 * gap,    i + 2 * width + 2 * gap],
                [text_y,    text_y * 1.03,  text_y * 1.03,              text_y],
                color='black',
                lw=1.5)
            ax.text(
                i + width + gap, text_y * 1.035,
                # f"mean={row['M2']}\nstd={row['STD2']}\np{'' if '<' in str(row['S2']) else '='}{row['S2']}",
                f"p{'' if '<' in str(row['S2']) else '='}{row['S2']}",
                ha='center', va='bottom')

    ax.set_xticks(ticks=xticks, labels=xlabels, rotation=45)

    for x, text in group_labels:
        ax.text(x,  options['max_tick'] + 23 - 1, text, ha='center', va='top', fontsize=15)

    plt.tight_layout()
    plt.savefig(file_name, dpi=300)


def is_significant(p_value):
    if '<' in str(p_value):
        return True

    if float(p_value) < 0.05:
        return True

    return False


# def draw_compare(file_name, table):

#     fig, ax = plt.subplots(figsize=(15, 8))
#     ax.set_ylim(50, 115)
#     ax.set_yticks(range(50, 105, 10))
#     ax.set_xlim(-0.2, 3.8)
#     ax.set_ylabel('Percentage', fontsize=14)
#     ax.tick_params(axis='both', labelsize=12)

#     width = 0.27
#     gap = 0.02
#     xticks = []
#     xlabels = []
#     group_labels = []

#     for i, group_name in enumerate(['Accuracy', 'Precision', 'Recall', 'F1']):
#         for j, label in enumerate(['GPT-4', 'Llama 3.1 70B', 'Llama 3.1 8B']):
#             row_id = i * 3 + j
#             row = table[row_id]

#             x = i + j * width + j * gap

#             ax.bar(
#                 x - width / 4, row['Base'], width / 2, color='#ADD8E6')
#             ax.text(
#                 x - width / 4, row['Base'] * 1.02, f"{row['Base']:.0f}%", ha='center', va='bottom')

#             ax.bar(
#                 x + width / 4, row['FT'], width / 2, color='#4682B4')
#             ax.text(
#                 x + width / 4, row['FT'] * 1.02, f"{row['FT']:.0f}%", ha='center', va='bottom')

#             xticks.append(x)
#             xlabels.append(label)

#             if is_significant(row['S']):
#                 ax.plot(
#                     [x - width / 2, x - width / 2,  x + width / 2,  x + width / 2],
#                     [105,           105 * 1.03,     105 * 1.03,     105],
#                     color='black',
#                     lw=1.5)
#                 ax.text(
#                     x, 105 * 1.035,
#                     f"mean={row['M']}\nstd={row['STD']}\np{'' if '<' in str(row['S']) else '='}{row['S']}",
#                     ha='center', va='bottom')

#         group_labels.append(
#             (
#                 i + width + gap,
#                 group_name
#             )
#         )

#     ax.set_xticks(ticks=xticks, labels=xlabels, rotation=45)

#     for x, text in group_labels:
#         ax.text(x,  114, text, ha='center', va='top', fontsize=15)

#     plt.tight_layout()
#     # plt.show()
#     plt.savefig(file_name, dpi=300)


# def draw_compare2(file_name, table):

#     fig, ax = plt.subplots(figsize=(15, 8))
#     ax.set_ylim(-10, 30)
#     ax.set_yticks(range(-10, 30, 5))
#     ax.set_xlim(-0.2, 3.8)
#     ax.set_ylabel('Percent difference', fontsize=14)
#     ax.tick_params(axis='both', labelsize=12)

#     width = 0.3
#     gap = 0.02
#     xticks = []
#     xlabels = []
#     group_labels = []

#     for i, group_name in enumerate(['Accuracy', 'Precision', 'Recall', 'F1']):
#         for j, label in enumerate(['GPT-4', 'Llama 3.1 70B', 'Llama 3.1 8B']):
#             row_id = i * 3 + j
#             row = table[row_id]

#             x = i + j * width + j * gap

#             ax.bar(
#                 x, row['D'], width, color='#D3D3D3')
#             ax.text(
#                 x, ((row['D'] * 1.02) if float(row['D']) > 0 else (row['D'] - 1)),
#                 f"{row['D']:.0f}%", ha='center', va='bottom')

#             xticks.append(x)
#             xlabels.append(label)

#             if is_significant(row['S']):
#                 ax.text(
#                     x, 23 * 1.035,
#                     f"mean={row['M']}\nstd={row['STD']}\np{'' if '<' in str(row['S']) else '='}{row['S']}",
#                     ha='center', va='bottom')

#         group_labels.append(
#             (
#                 i + width + gap,
#                 group_name
#             )
#         )

#     ax.set_xticks(ticks=xticks, labels=xlabels, rotation=45)

#     for x, text in group_labels:
#         ax.text(x,  29, text, ha='center', va='top', fontsize=15)

#     plt.tight_layout()
#     # plt.show()
#     plt.savefig(file_name, dpi=300)


def draw_compare3(file_name, table):

    # sort table by category (Accuracy, Precision, Recall, F1) then by model (GPT-4, Llama 3.1 70B, Llama 3.1 8B)
    if table:
        keys = list(table[0].keys())
        cat_key = next((k for k in keys if k.lower() in ('category', 'cat')), None)
        model_key = next((k for k in keys if k.lower() == 'model'), None)

        cat_order = ['Accuracy', 'Precision', 'Recall', 'F1']
        model_order = ['GPT-4', 'Llama 3.1 70B', 'Llama 3.1 8B']

        def cat_index(row):
            if not cat_key:
                return 0
            v = str(row.get(cat_key, '')).strip()
            for i, c in enumerate(cat_order):
                if v.lower() == c.lower():
                    return i
            return len(cat_order)

        def model_index(row):
            if not model_key:
                return 0
            v = str(row.get(model_key, '')).lower()
            if 'gpt' in v:
                return 0
            if '70' in v or '70b' in v:
                return 1
            if '8' in v or '8b' in v:
                return 2
            for i, m in enumerate(model_order):
                if v == m.lower():
                    return i
            return len(model_order)

        table.sort(key=lambda r: (cat_index(r), model_index(r)))


    fig, ax = plt.subplots(figsize=(15, 8))
    ax.set_ylim(-13, 30)
    ax.set_yticks(range(-10, 30, 5))
    ax.set_xlim(-0.2, 3.8)
    ax.set_ylabel('Percent difference', fontsize=14)
    ax.tick_params(axis='both', labelsize=12)

    width = 0.3
    gap = 0.02
    xticks = []
    xlabels = []
    group_labels = []

    for i, group_name in enumerate(['Accuracy', 'Precision', 'Recall', 'F1']):
        for j, label in enumerate(['GPT-4', 'Llama 3.1 70B', 'Llama 3.1 8B']):
            row_id = i * 3 + j
            row = table[row_id]

            x = i + j * width + j * gap

            ax.bar(
                x, (row['D'] if float(row['D']) else 0.3), width, color='#D3D3D3')
            ax.text(
                x - width / 4,
                ((row['D'] * 1.02) if float(row['D']) > 0 else (row['D'] - 2)),
                f"{row['D']:.0f}%", ha='center', va='bottom')

            y2 = ((row['D'] + row['STD'] / sqrt(16)) if row['D'] > 0 else (row['D'] - row['STD'] / sqrt(16)))
            ax.plot(
                [x,         x],
                [row['D'],  y2],
                color='black',
                lw=1.5)

            ax.plot(
                [x - width / 8, x + width / 8],
                [y2,            y2],
                color='black',
                lw=1.5)

            xticks.append(x)
            xlabels.append(label)

            if is_significant(row['S']):
                ax.text(
                    x, (y2 * 1.05 if float(row['D']) > 0 else y2 * 1.2),
                    # f"mean={row['M']}\nstd={row['STD']}\np{'' if '<' in str(row['S']) else '='}{row['S']}",
                    f"p{'' if '<' in str(row['S']) else '='}{row['S']}",
                    ha='center', va='bottom')

        group_labels.append(
            (
                i + width + gap,
                group_name
            )
        )

    ax.set_xticks(ticks=xticks, labels=xlabels, rotation=45)

    for x, text in group_labels:
        ax.text(x,  29, text, ha='center', va='top', fontsize=15)

    plt.tight_layout()
    # plt.show()
    plt.savefig(file_name, dpi=300)


def work():
    baseline = load_excel('Data Nov 22.xlsx', 'baseline')
    draw_baseline('baseline.png', baseline, '#ADD8E6', max_tick=87)

    compare = load_excel('Data Nov 22.xlsx', 'compare')
    draw_compare3('Compare.png', compare)

    ft = load_excel('Data Nov 22.xlsx', 'FT')
    draw_baseline('fine-tuned.png', ft, '#4682B4', max_tick=97)


if __name__ == '__main__':
    work()
